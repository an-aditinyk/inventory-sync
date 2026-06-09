"""Bridges between raw inputs (an .xlsx upload, a Shopify connection) and the
engine's readers. Keeps file/HTTP concerns out of the pure engine.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from .engine import ExcelReader, ShopifyReader, SourceItem


def parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Turn xlsx bytes into a list of header->value row dicts (first sheet).

    The first row is treated as the header. Blank trailing rows are skipped.
    """

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []

    rows: list[dict[str, Any]] = []
    for raw in rows_iter:
        if raw is None or all(cell is None for cell in raw):
            continue
        rows.append({headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)})
    workbook.close()
    return rows


def read_excel_source(content: bytes, column_map: dict[str, str]) -> list[SourceItem]:
    rows = parse_xlsx(content)
    return ExcelReader(column_map).read(rows)


def read_shopify_source(variants: list[dict[str, Any]]) -> list[SourceItem]:
    """Normalize already-fetched Shopify variants.

    For local dev the API accepts variants as JSON. A real Shopify pull (OAuth
    token → GraphQL inventory query → summed-per-SKU variants) produces the same
    shape and then flows through here unchanged.
    """

    return ShopifyReader().read(variants)
